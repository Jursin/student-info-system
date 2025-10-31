import io
import os
from datetime import datetime
from flask import Blueprint, render_template, request, redirect, url_for, flash, session, send_file, current_app
from sqlalchemy import or_
from ..extensions import db
from ..models import Student, Admin
from openpyxl import Workbook
from ..forms import BulkImportForm, AdminLoginForm, StudentEditForm, StudentCreateForm
from werkzeug.utils import secure_filename
import csv


bp = Blueprint("admin", __name__)


def _is_logged_in() -> bool:
    return bool(session.get("admin_logged_in"))


@bp.route("/login", methods=["GET", "POST"])
def login():
    form = AdminLoginForm()
    if form.validate_on_submit():
        admin = Admin.query.filter_by(username=form.username.data.strip()).first()
        if admin and admin.check_password(form.password.data):
            session["admin_logged_in"] = True
            session["admin_username"] = admin.username
            return redirect(url_for("admin.list_students"))
        flash("用户名或密码错误", "danger")
    return render_template("admin_login.html", form=form)


@bp.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("admin.login"))


@bp.route("/students")
def list_students():
    if not _is_logged_in():
        return redirect(url_for("admin.login"))

    q = request.args.get("q", "").strip()
    genders = [g.strip() for g in request.args.getlist("gender") if g.strip()]
    majors = [m.strip() for m in request.args.getlist("major") if m.strip()]
    clazz_param = request.args.get("clazz", "").strip()
    clazz_list = [c for c in [x.strip() for x in clazz_param.split(",")] if c]
    page = max(int(request.args.get("page", 1) or 1), 1)
    per_page = 40

    query = Student.query
    if q:
        like = f"%{q}%"
        query = query.filter(
            or_(
                Student.name.like(like),
                Student.student_id.like(like),
                Student.id_card.like(like),
                Student.phone.like(like),
                Student.major.like(like),
                Student.clazz.like(like),
            )
        )
    if genders:
        query = query.filter(Student.gender.in_(genders))
    if majors:
        query = query.filter(Student.major.in_(majors))
    if clazz_list:
        query = query.filter(Student.clazz.in_(clazz_list))

    pagination = query.order_by(Student.student_id.asc()).paginate(page=page, per_page=per_page, error_out=False)
    import_form = BulkImportForm()
    current_gender = genders[0] if genders else ""
    current_major = majors[0] if majors else ""
    return render_template(
        "admin_list.html",
        students=pagination.items,
        q=q,
        gender=current_gender,
        major=current_major,
        clazz=clazz_param,
        import_form=import_form,
        pagination=pagination,
    )


@bp.route("/export")
def export_excel():
    import io
    if not _is_logged_in():
        return redirect(url_for("admin.login"))

    q = request.args.get("q", "").strip()
    gender = request.args.get("gender", "").strip()
    major = request.args.get("major", "").strip()
    clazz = request.args.get("clazz", "").strip()
    fmt = (request.args.get("format", "xlsx") or "xlsx").lower()

    all_fields = [
        "student_id", "name", "gender", "hometown", "id_card", "phone", "major", "clazz", "admin_class", "created_at", "updated_at"
    ]
    req_fields = request.args.getlist("fields")
    if not req_fields:
        fields = all_fields
    else:
        # 只用 intersection 部分, 且用页面顺序排序
        fields = [f for f in all_fields if f in req_fields]

    student_ids = request.args.getlist("student_ids")
    field_map = {
        "student_id": "学号",
        "name": "姓名",
        "gender": "性别",
        "hometown": "籍贯",
        "id_card": "身份证号码",
        "phone": "电话号码",
        "major": "专业",
        "clazz": "班级",
        "admin_class": "行政班级",
        "created_at": "创建时间",
        "updated_at": "更新时间"
    }
    headers = [field_map.get(field, field) for field in fields]
    if student_ids:
        query = Student.query.filter(Student.id.in_(student_ids))
        students = query.order_by(Student.updated_at.desc()).all()
    else:
        students = []
    if fmt == "csv":
        import csv
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(headers)
        for s in students:
            row = []
            for field in fields:
                if field == "created_at":
                    row.append(s.created_at.strftime("%Y-%m-%d %H:%M:%S") if s.created_at else "")
                elif field == "updated_at":
                    row.append(s.updated_at.strftime("%Y-%m-%d %H:%M:%S") if s.updated_at else "")
                else:
                    row.append(getattr(s, field, ""))
            w.writerow(row)
        data = io.BytesIO(buf.getvalue().encode("utf-8-sig"))
        filename = f"students_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        return send_file(data, as_attachment=True, download_name=filename, mimetype="text/csv")
    else:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "学生信息"
        ws.append(headers)
        for s in students:
            row = []
            for field in fields:
                if field == "created_at":
                    row.append(s.created_at.strftime("%Y-%m-%d %H:%M:%S") if s.created_at else "")
                elif field == "updated_at":
                    row.append(s.updated_at.strftime("%Y-%m-%d %H:%M:%S") if s.updated_at else "")
                else:
                    row.append(getattr(s, field, ""))
            ws.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"students_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(buf, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@bp.route("/import", methods=["POST"])
def import_bulk():
    if not _is_logged_in():
        return redirect(url_for("admin.login"))
    form = BulkImportForm()
    if not form.validate_on_submit():
        flash("请选择文件", "danger")
        return redirect(url_for("admin.list_students"))

    file = request.files.get("file")
    if not file or file.filename == "":
        flash("未选择文件", "danger")
        return redirect(url_for("admin.list_students"))

    filename = secure_filename(file.filename)
    ext = os.path.splitext(filename)[1].lower()
    current_app.logger.info("[IMPORT] Start import: filename=%s, ext=%s", filename, ext)

    # 流式读取，避免占用过多内存
    try:
        headers = []
        row_iter = None
        is_excel = False
        MAX_ROWS = 20000
        if ext in [".xlsx", ".xlsm", ".xltx", ".xltm"]:
            from openpyxl import load_workbook
            wb = load_workbook(file, data_only=True, read_only=True)
            ws = wb.active
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [ (c or "").strip() if isinstance(c, str) else ("" if c is None else str(c).strip()) for c in header_row ]
            current_app.logger.info("[IMPORT] Detected headers (xlsx): %s", headers)
            row_iter = ws.iter_rows(min_row=2, values_only=True)
            is_excel = True
        elif ext in [".csv", ".txt"]:
            import io as _io
            file.stream.seek(0)
            text_stream = _io.TextIOWrapper(file.stream, encoding="utf-8-sig", newline="")
            reader = csv.reader(text_stream)
            headers = next(reader, [])
            current_app.logger.info("[IMPORT] Detected headers (csv): %s", headers)
            row_iter = reader
        else:
            flash("不支持的文件格式", "danger")
            current_app.logger.warning("[IMPORT] Unsupported file extension: %s", ext)
            return redirect(url_for("admin.list_students"))
    except Exception as e:
        flash(f"文件读取失败: {e}", "danger")
        current_app.logger.exception("[IMPORT] Failed to read file: %s", e)
        return redirect(url_for("admin.list_students"))

    # 允许缺列为空，多余列忽略；忽略“行政班级”列，自动生成
    # 支持表头中文：姓名、性别、学号、身份证号码、电话号码/电话、专业、班级、籍贯、行政班级（忽略）
    key_alias = {
        "姓名": "name",
        "性别": "gender",
        "学号": "student_id",
        "身份证": "id_card",
        "身份证号码": "id_card",
        "电话": "phone",
        "电话号码": "phone",
        "专业": "major",
        "班级": "clazz",
        "籍贯": "hometown",
        "行政班级": "admin_class_ignored",
    }
    # 规范化表头 -> 字段名（去空白、大小写、BOM、常见中英对照）
    def _norm_str(s):
        if s is None:
            return ""
        s = str(s).strip().lstrip("\ufeff").lower()
        return s
    header_alias_ext = {
        "name": "name",
        "gender": "gender",
        "student_id": "student_id",
        "studentid": "student_id",
        "id": "student_id",
        "id_card": "id_card",
        "idcard": "id_card",
        "identity": "id_card",
        "phone": "phone",
        "mobile": "phone",
        "tel": "phone",
        "major": "major",
        "class": "clazz",
        "clazz": "clazz",
        "class_no": "clazz",
        "hometown": "hometown",
        "admin_class": "admin_class_ignored",
        "adminclass": "admin_class_ignored",
    }
    norm_headers = []
    for h in headers:
        raw = _norm_str(h)
        # 先中文别名，再英文别名
        mapped = key_alias.get(h if isinstance(h, str) else h, "")
        if not mapped:
            mapped = header_alias_ext.get(raw, "")
        norm_headers.append(mapped)
    current_app.logger.info("[IMPORT] Normalized headers: %s", norm_headers)
    if not any(norm_headers):
        # 无法识别表头时，按默认顺序做位置推断；使用 headers 的长度作为基准
        length = len(headers)
        first_vals = []
        default_order = ["name", "gender", "student_id", "id_card", "phone", "major", "clazz", "hometown"]
        lead_empty = 1 if (first_vals and str(first_vals[0]).strip().isdigit() and length >= 2) else 0
        norm_headers = ([""] * lead_empty) + default_order
        norm_headers = norm_headers[:length]
        current_app.logger.warning("[IMPORT] Using positional header mapping: %s", norm_headers)

    inserted, updated, failed = 0, 0, 0
    processed = 0
    # 记录本次导入过程中新增但尚未提交的 student_id -> Student 对象，避免同一文件中重复学号导致唯一键冲突
    pending_new_by_sid = {}
    BATCH_SIZE = 500
    for idx, row in enumerate(row_iter, start=2):  # 数据从第2行开始
        try:
            values = {}
            # Excel: row 是值列表；CSV: row 是字符串列表
            if is_excel:
                cells = list(row)
                if not any(cells):
                    continue
                for cidx, field in enumerate(norm_headers):
                    if not field or field == "admin_class_ignored":
                        continue
                    if cidx >= len(cells):
                        values[field] = ""
                        continue
                    cv = cells[cidx]
                    if field in ("student_id", "id_card", "phone", "clazz"):
                        # 强制转字符串并做必要补零
                        if cv is None:
                            s = ""
                        elif isinstance(cv, (int, float)):
                            # 避免 1.0 之类格式
                            iv = int(cv)
                            s = str(iv)
                        else:
                            s = str(cv).strip()
                        if field == "student_id":
                            s = s.zfill(10) if s else s
                        elif field == "id_card":
                            s = s.zfill(18) if s and len(s) < 18 and s.isdigit() else s
                        elif field == "phone":
                            s = s.zfill(11) if s and len(s) < 11 and s.isdigit() else s
                        elif field == "clazz":
                            s = s.zfill(5) if s and len(s) < 5 and s.isdigit() else s
                        values[field] = s
                    else:
                        values[field] = "" if cv is None else str(cv).strip()
            else:  # CSV/TXT
                cells = ["" if x is None else str(x).strip() for x in row]
                if not any(cells):
                    continue
                for cidx, field in enumerate(norm_headers):
                    if not field or field == "admin_class_ignored":
                        continue
                    v = cells[cidx] if cidx < len(cells) else ""
                    if field == "student_id":
                        v = v.zfill(10) if v and v.isdigit() and len(v) < 10 else v
                    elif field == "id_card":
                        v = v.zfill(18) if v and v.isdigit() and len(v) < 18 else v
                    elif field == "phone":
                        v = v.zfill(11) if v and v.isdigit() and len(v) < 11 else v
                    elif field == "clazz":
                        v = v.zfill(5) if v and v.isdigit() and len(v) < 5 else v
                    values[field] = v

            name = values.get("name", "")
            gender = values.get("gender", "")
            student_id = values.get("student_id", "")
            id_card = values.get("id_card", "")
            phone = values.get("phone", "")
            major = values.get("major", "")
            clazz = values.get("clazz", "")
            hometown = values.get("hometown", "")

            # 仅要求识别标识和基本姓名，其余字段若缺失则留空
            if not all([student_id, name]):
                failed += 1
                current_app.logger.warning(
                    "[IMPORT] Row %d failed required check: values=%s", idx, values
                )
                continue

            # 先检查本批次新增缓存，再查询数据库；使用 no_autoflush 防止查询触发未提交对象的提前 flush
            stu = pending_new_by_sid.get(student_id)
            if stu is None:
                with db.session.no_autoflush:
                    stu = Student.query.filter_by(student_id=student_id).first()
            if stu:
                stu.name = name
                stu.gender = gender
                stu.id_card = id_card
                stu.phone = phone
                stu.major = major
                stu.clazz = clazz
                stu.hometown = hometown or ""
                stu.update_admin_class()
                updated += 1
                current_app.logger.debug("[IMPORT] Row %d -> update student_id=%s", idx, student_id)
            else:
                new_stu = Student(
                    name=name,
                    gender=gender,
                    student_id=student_id,
                    id_card=id_card,
                    phone=phone,
                    major=major,
                    clazz=clazz,
                    hometown=hometown or "",
                )
                new_stu.update_admin_class()
                db.session.add(new_stu)
                pending_new_by_sid[student_id] = new_stu
                inserted += 1
                current_app.logger.debug("[IMPORT] Row %d -> insert student_id=%s", idx, student_id)
            processed += 1
            if processed % BATCH_SIZE == 0:
                try:
                    db.session.flush()
                except Exception as _e:
                    current_app.logger.exception("[IMPORT] Flush failed at row %d", idx)
        except Exception:
            failed += 1
            current_app.logger.exception("[IMPORT] Row %d processing error", idx)
        if processed >= MAX_ROWS:
            current_app.logger.warning("[IMPORT] Reached MAX_ROWS limit: %d", MAX_ROWS)
            break
    try:
        db.session.commit()
        current_app.logger.info("[IMPORT] Done. inserted=%d, updated=%d, failed=%d", inserted, updated, failed)
    except Exception as e:
        db.session.rollback()
        flash(f"导入提交失败: {e}", "danger")
        current_app.logger.exception("[IMPORT] Commit failed: %s", e)
        return redirect(url_for("admin.list_students"))

    flash(f"导入完成：新增 {inserted} 条，更新 {updated} 条，失败 {failed} 条", "success")
    return redirect(url_for("admin.list_students"))


@bp.route("/students/delete", methods=["POST"])
def delete_students():
    if not _is_logged_in():
        return redirect(url_for("admin.login"))
    
    student_ids = request.form.getlist("student_ids")
    if not student_ids:
        flash("请选择要删除的学生", "warning")
        return redirect(url_for("admin.list_students"))
    
    try:
        deleted_count = Student.query.filter(Student.id.in_(student_ids)).delete(synchronize_session=False)
        db.session.commit()
        flash(f"已删除 {deleted_count} 条记录", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"删除失败: {e}", "danger")
    
    return redirect(url_for("admin.list_students"))


@bp.route("/students/<int:student_pk>/edit", methods=["GET", "POST"])
def edit_student(student_pk: int):
    if not _is_logged_in():
        return redirect(url_for("admin.login"))
    stu = Student.query.get_or_404(student_pk)
    # 管理端允许修改学号，使用创建表单以包含 student_id 字段
    form = StudentCreateForm(obj=stu)
    messages = []
    if form.validate_on_submit():
        new_sid = form.student_id.data.strip()
        if new_sid != stu.student_id:
            exists = Student.query.filter_by(student_id=new_sid).first()
            if exists:
                messages.append(("danger", "学号已存在"))
                return render_template("admin_edit.html", form=form, stu=stu, messages=messages)
            stu.student_id = new_sid
        stu.name = form.name.data.strip()
        stu.gender = form.gender.data.strip()
        stu.id_card = form.id_card.data.strip()
        stu.phone = form.phone.data.strip()
        stu.major = form.major.data.strip()
        stu.clazz = form.clazz.data.strip()
        stu.hometown = form.hometown.data.strip()
        stu.update_admin_class()
        try:
            db.session.commit()
            messages.append(("success", "已保存"))
        except Exception as e:
            db.session.rollback()
            messages.append(("danger", "保存失败: %s" % e))
    return render_template("admin_edit.html", form=form, stu=stu, messages=messages)


@bp.route("/students/create", methods=["GET", "POST"])
def create_student():
    if not _is_logged_in():
        return redirect(url_for("admin.login"))
    form = StudentCreateForm()
    if form.validate_on_submit():
        # 学号唯一创建
        exists = Student.query.filter_by(student_id=form.student_id.data.strip()).first()
        if exists:
            flash("学号已存在", "danger")
            return render_template("admin_create.html", form=form)
        stu = Student(
            name=form.name.data.strip(),
            gender=form.gender.data.strip(),
            student_id=form.student_id.data.strip(),
            id_card=form.id_card.data.strip(),
            phone=form.phone.data.strip(),
            major=form.major.data.strip(),
            clazz=form.clazz.data.strip(),
            hometown=form.hometown.data.strip(),
        )
        stu.update_admin_class()
        try:
            db.session.add(stu)
            db.session.commit()
            flash("新增成功", "success")
            return redirect(url_for("admin.list_students"))
        except Exception as e:
            db.session.rollback()
            flash(f"新增失败: {e}", "danger")
    # 校验失败或首次进入
    return render_template("admin_create.html", form=form)

